"""
theme.py — Mission 26004 / Que Choisir Ensemble
Charte graphique, formatage des valeurs et composants d'affichage du dashboard.
Seul module autorise a definir des couleurs : aucun hexadecimal ailleurs.
"""

import io
import re

import pandas as pd
import streamlit as st

# ---------- charte QCE ----------
BLEU = "#004F9F"
BLEU_FONCE = "#00274F"
BLEU_INTER = "#003A77"
ROUGE = "#E30613"
JAUNE = "#FFF03C"
VERT = "#3FA535"
GRIS = "#E8E8E8"
BLANC = "#FFFFFF"
GRIS_TEXTE = "#5A6672"
JAUNE_FOND = "#FFFBE0"
TRANSPARENT = "rgba(0,0,0,0)"

ECHELLE_BLEUE = ["#E8EFF7", "#C2D6EB", "#8FB4DA", "#5A8FC7", "#2E6EB0", "#004F9F", "#00274F"]
ECHELLE_BLEUE_INV = list(reversed(ECHELLE_BLEUE))
SEQUENCE_CATEGORIES = [BLEU, BLEU_INTER, BLEU_FONCE, VERT, JAUNE, ROUGE]

CATEGORIES_EFFORT = {
    "Accessible": ECHELLE_BLEUE[1],
    "Modéré": ECHELLE_BLEUE[3],
    "Effort important": BLEU,
    "Insoutenable": ROUGE,
}

POLICE = "Source Sans Pro, Segoe UI, Helvetica, Arial, sans-serif"


# ---------- formatage : point d'entree unique ----------
def fmt(valeur, genre="entier", unite=True):
    """Applique les arrondis d'affichage du rapport. genre : taux, taux0, taux2,
    euro, entier, surface, ratio, ratio1, decimal2."""
    if valeur is None:
        return "n.d."
    try:
        x = float(valeur)
    except (TypeError, ValueError):
        return "n.d."
    if x != x:
        return "n.d."

    if genre == "taux":
        s = f"{x:.1f}".replace(".", ",")
        return f"{s} %" if unite else s
    if genre == "taux0":
        s = f"{x:.0f}"
        return f"{s} %" if unite else s
    if genre == "taux2":
        s = f"{x:.2f}".replace(".", ",")
        return f"{s} %" if unite else s
    if genre == "surface":
        s = f"{x:.1f}".replace(".", ",")
        return f"{s} m²" if unite else s
    if genre == "ratio":
        s = f"{x:.2f}".replace(".", ",")
        return s
    if genre == "ratio1":
        s = f"{x:.1f}".replace(".", ",")
        return s
    if genre == "decimal2":
        return f"{x:,.2f}".replace(",", " ").replace(".", ",")
    if genre == "euro":
        s = f"{x:,.0f}".replace(",", " ")
        return f"{s} €" if unite else s
    s = f"{x:,.0f}".replace(",", " ")
    return s


def fmt_delta(valeur, genre="taux"):
    if valeur is None or valeur != valeur:
        return None
    signe = "+" if valeur > 0 else ""
    return f"{signe}{fmt(valeur, genre)}"


# ---------- mise en page ----------
def configurer_page(titre):
    st.set_page_config(page_title=f"{titre} — QCE", page_icon="🏠",
                       layout="wide", initial_sidebar_state="expanded")
    st.markdown(f"""
        <style>
        .stApp {{ background: {BLANC}; }}
        h1, h2, h3 {{ color: {BLEU_FONCE}; font-family: {POLICE}; }}
        [data-testid="stSidebar"] {{ background: {GRIS}; }}
        [data-testid="stSidebar"] * {{ color: {BLEU_FONCE}; }}
        .bandeau {{ background: {BLEU}; color: {BLANC}; padding: 1.1rem 1.4rem;
                    border-radius: 6px; margin-bottom: 1.1rem; }}
        .bandeau h1 {{ color: {BLANC}; margin: 0; font-size: 1.55rem; }}
        .bandeau p {{ color: {BLANC}; margin: .35rem 0 0; opacity: .92; font-size: .95rem; }}
        .kpi {{ background: {BLANC}; border: 1px solid {GRIS}; border-left: 5px solid {BLEU};
                border-radius: 5px; padding: .85rem 1rem; height: 100%; }}
        .kpi.critique {{ border-left-color: {ROUGE}; }}
        .kpi .libelle {{ color: {GRIS_TEXTE}; font-size: .78rem; text-transform: uppercase;
                         letter-spacing: .04em; }}
        .kpi .valeur {{ color: {BLEU_FONCE}; font-size: 1.85rem; font-weight: 700;
                        line-height: 1.15; margin: .18rem 0; }}
        .kpi.critique .valeur {{ color: {ROUGE}; }}
        .kpi .note {{ color: {GRIS_TEXTE}; font-size: .78rem; }}
        .perimetre {{ background: {JAUNE_FOND}; border-left: 4px solid {JAUNE};
                      padding: .6rem .9rem; border-radius: 4px; font-size: .87rem;
                      color: {BLEU_FONCE}; margin: .5rem 0 .9rem; }}
        .source {{ color: {GRIS_TEXTE}; font-size: .78rem; border-top: 1px solid {GRIS};
                   padding-top: .5rem; margin-top: .8rem; }}
        </style>
    """, unsafe_allow_html=True)


def bandeau(titre, sous_titre):
    st.markdown(f'<div class="bandeau"><h1>{titre}</h1><p>{sous_titre}</p></div>',
                unsafe_allow_html=True)


def kpi(libelle, valeur, note="", critique=False):
    classe = "kpi critique" if critique else "kpi"
    st.markdown(
        f'<div class="{classe}"><div class="libelle">{libelle}</div>'
        f'<div class="valeur">{valeur}</div><div class="note">{note}</div></div>',
        unsafe_allow_html=True)


def perimetre_modeste(n_communes=5163, inline=False):
    """Rappel obligatoire : le menage modeste ne porte pas sur le territoire entier."""
    texte = (f"Le ménage modeste (1er décile) porte sur "
             f"<b>{fmt(n_communes)} communes</b> — celles pour lesquelles l'INSEE publie le "
             f"premier décile de revenu — et non sur l'ensemble du territoire.")
    if inline:
        return texte
    st.markdown(
        f'<div class="perimetre">⚠️ {texte}</div>', unsafe_allow_html=True)


def source(texte):
    st.markdown(f'<div class="source">{texte}</div>', unsafe_allow_html=True)


def mise_en_forme_graphique(fig, hauteur=420, titre=""):
    fig.update_layout(
        height=hauteur,
        font=dict(family=POLICE, color=BLEU_FONCE, size=13),
        paper_bgcolor=BLANC, plot_bgcolor=BLANC,
        margin=dict(l=10, r=10, t=86 if titre else 34, b=10),
        title=dict(text=titre, font=dict(color=BLEU_FONCE, size=15),
                   x=0, xanchor="left", y=1, yanchor="top", yref="container",
                   pad=dict(t=12, l=6)),
        legend=dict(orientation="h", yanchor="bottom", y=1.015, xanchor="left", x=0,
                    font=dict(size=11)),
        hoverlabel=dict(bgcolor=BLANC, bordercolor=BLEU, font_color=BLEU_FONCE),
    )
    fig.update_xaxes(gridcolor=GRIS, zerolinecolor=GRIS, linecolor=GRIS)
    fig.update_yaxes(gridcolor=GRIS, zerolinecolor=GRIS, linecolor=GRIS)
    return fig


# ---------- exports ----------
# Les deux formats exportent le MEME DataFrame, colonnes techniques comprises.
# Seul l'encodage du fichier change ; aucune valeur n'est reformatee.

# Excel interdit  : \ / ? * [ ]  dans un nom de feuille, et le limite a 31 signes.
_INTERDITS_FEUILLE = re.compile(r"[:\\/?*\[\]]")


def nom_feuille(libelle):
    """Rend un libelle utilisable comme nom d'onglet Excel."""
    propre = _INTERDITS_FEUILLE.sub(" ", str(libelle)).strip()
    return (propre[:31] or "Donnees").strip("'")


def _classeur_xlsx(df, feuille):
    """Serialise df en .xlsx : en-tetes en gras, filtres, volets figes, largeurs.

    Les valeurs sont ecrites telles quelles — les nombres restent des nombres,
    Excel les affiche selon la locale du poste.
    """
    # Import local : une absence d'openpyxl ne doit pas empecher toute
    # l'application de demarrer, seulement ce bouton de repondre.
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    feuille = nom_feuille(feuille)
    tampon = io.BytesIO()
    with pd.ExcelWriter(tampon, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=feuille, index=False)
        ws = writer.sheets[feuille]

        for cellule in ws[1]:
            cellule.font = Font(bold=True, color="FFFFFF")
            cellule.fill = PatternFill("solid", fgColor=BLEU.lstrip("#"))
            cellule.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # largeur : le plus long entre l'en-tete et les 200 premieres valeurs,
        # borne a [10, 42] pour qu'aucune colonne ne devienne illisible.
        for i, colonne in enumerate(df.columns, start=1):
            valeurs = df[colonne].head(200)
            large = max([len(str(colonne))] + [len(_texte(v)) for v in valeurs] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(max(large + 3, 10), 42)

    return tampon.getvalue()


def _texte(v):
    if v is None or (isinstance(v, float) and v != v):
        return ""
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)


def bouton_csv(df, nom_fichier, libelle="Télécharger (CSV)"):
    """CSV lisible par un Excel francais : point-virgule, virgule decimale, BOM."""
    st.download_button(
        label=libelle,
        data=df.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig"),
        file_name=nom_fichier,
        mime="text/csv",
        width="stretch",
    )


def bouton_excel(df, nom_fichier, feuille="Donnees", libelle="Télécharger (Excel)"):
    """Jumeau de bouton_csv, au format .xlsx."""
    st.download_button(
        label=libelle,
        data=_classeur_xlsx(df, feuille),
        file_name=nom_fichier,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )


def boutons_export(df, base, feuille):
    """Les deux boutons cote a cote. `base` est le nom de fichier sans extension."""
    g, d = st.columns(2)
    with g:
        bouton_csv(df, f"{base}.csv")
    with d:
        bouton_excel(df, f"{base}.xlsx", feuille)
